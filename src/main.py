import os
import csv
from openpyxl import Workbook, load_workbook

INPUT_DIR = "input"
OUTPUT_DIR = "output"
OUTPUT_FILE = "merged.xlsx"

# Columns to adjust according to your file
NAME_COL = 0       # index of the "Name" column
EMAIL_COL = 1      # index of the "Email" column
PRICE_COL = 2      # index of the "Price" column
PHONE_COL = 3      # index of the "Phone" column

os.makedirs(OUTPUT_DIR, exist_ok=True)
output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)

def merge_all_files():
    wb_out = Workbook()
    ws_out = wb_out.active

    files_found = False

    for filename in os.listdir(INPUT_DIR):
        path = os.path.join(INPUT_DIR, filename)

        # CSV files
        if filename.lower().endswith(".csv"):
            files_found = True
            with open(path, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    ws_out.append(row) 

        # Excel files
        elif filename.lower().endswith((".xls", ".xlsx")):
            files_found = True
            wb_in = load_workbook(path, read_only=True)
            sheet_in = wb_in.active
            for row in sheet_in.iter_rows(values_only=True):
                ws_out.append(list(row))

    if not files_found:
        raise ValueError("No CSV or Excel files found in the folder.")

    return wb_out, ws_out

def standardize_data(ws):
    for row in iter_rows(min_row = 2):
        if row[NAME_COL].value:
            row[NAME_COL].value = str(row[NAME_COL].value).upper()

        if row[PRICE_COL].value:
            try:
                row[PRICE_COL].value = float(str(row[PRICE_COL].value).replace(",", ".").replace("€", "").strip())
            except:
                row[PRICE_COL].value = 0.0

        if row[phone].value:
            row[NAME_COL].value = str(row[NAME_COL].value).upper()
        
        if row[PHONE_COL].value:
            num = ''.join(filter(str.isdigit, str(row[PHONE_COL].value)))
            if not num.startswith("32"):  # Belgium country code
                num = "32" + num
            row[PHONE_COL].value = "+" + num

def remove_invalid_rows(ws):
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Delete row if Name or Email is missing
        if not row[NAME_COL].value or not row[EMAIL_COL].value:
            rows_to_delete.append(i)
    for i in reversed(rows_to_delete):
        ws.delete_rows(i)

def remove_duplicates(ws):
    seen = set()
    rows_to_delete = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Use Name + Email as duplicate key
        key = (row[NAME_COL].value, row[EMAIL_COL].value)
        if key in seen:
            rows_to_delete.append(i)
        else:
            seen.add(key)
    for i in reversed(rows_to_delete):
        ws.delete_rows(i)

def generate_summary(ws):
    # Total clients
    total_clients = ws.max_row - 1  # assuming first row is header
    total_revenue = 0
    for row in ws.iter_rows(min_row=2):
        try:
            total_revenue += float(row[PRICE_COL].value)
        except:
            continue
    print(f"Total clients: {total_clients}")
    print(f"Total revenue: {total_revenue} €")

# --- MAIN ---
wb_out, ws_out = merge_all_files()
standardize_data(ws_out)
remove_invalid_rows(ws_out)
remove_duplicates(ws_out)
generate_summary(ws_out)
wb_out.save(output_path)
print(f"Merging and cleaning completed → {output_path}")